from __future__ import annotations

from omni_eda import EDAConfig, OmniEDA
from omni_eda.export import export_all, export_csv_bundle, export_excel, export_json


def _run(basic_df, tmp_path):
    cfg = EDAConfig(verbose=False, output_dir=str(tmp_path / "out"), sample_for_plots=200, max_rows_for_expensive_ops=300)
    eda = OmniEDA(config=cfg)
    eda.run(basic_df)
    return eda.results, cfg


def test_export_json_is_valid(basic_df, tmp_path):
    import json

    results, cfg = _run(basic_df, tmp_path)
    path = export_json(results, tmp_path / "out.json")
    data = json.loads(path.read_text())
    assert data["shape"] == list(basic_df.shape)


def test_export_excel_has_sheets(basic_df, tmp_path):
    results, cfg = _run(basic_df, tmp_path)
    path = export_excel(results, tmp_path / "out.xlsx")
    assert path.exists()
    import openpyxl

    wb = openpyxl.load_workbook(path)
    assert "Overview" in wb.sheetnames
    assert "Statistics" in wb.sheetnames


def test_export_csv_bundle_creates_files(basic_df, tmp_path):
    results, cfg = _run(basic_df, tmp_path)
    written = export_csv_bundle(results, tmp_path / "csvs")
    assert written
    assert all(p.exists() for p in written)


def test_export_all_html_and_dashboard(basic_df, tmp_path):
    results, cfg = _run(basic_df, tmp_path)
    written = export_all(results, tmp_path / "exports", ["html", "dashboard"], cfg)
    assert "html" in written and written["html"].exists()
    assert "dashboard" in written and written["dashboard"].exists()
    dashboard_html = written["dashboard"].read_text()
    assert "Plotly" in dashboard_html
